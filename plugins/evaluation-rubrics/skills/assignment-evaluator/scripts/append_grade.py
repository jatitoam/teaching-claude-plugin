"""
append_grade.py — Assignment Evaluator output script
Appends one evaluated row to the grades xlsx, creating the file if it doesn't exist.

Usage:
    python append_grade.py <rubric_xlsx> <grades_xlsx> <row_json>

Arguments:
    rubric_xlsx   Path to the rubric .xlsx produced by rubric-creator
    grades_xlsx   Path to the grades output file (created if missing)
    row_json      Path to a JSON file with the evaluated row:

Row JSON structure:
{
  "name": "Group 1",
  "scores": {
    "Criterion Name 1": "Meets Expectations",
    "Criterion Name 2": "Partially Meets",
    "Criterion Name 3": "Does Not Meet"
  },
  "observations": {
    "Criterion Name 1": "",
    "Criterion Name 2": "only 2 out of 10 elements completed.",
    "Criterion Name 3": "missing price and link."
  },
  "penalties": {
    "Presentation": "met",
    "Attached file": {"level": "not_met", "note": "submitted a cloud link, not a file"}
  }
}

observations may also be a plain string (legacy format) — both are accepted.
When a dict is provided, non-empty values are compiled into a single string:
"Criterion Name 2: only 2 out of 10 elements completed. Criterion Name 3: missing price and link."

penalties is OPTIONAL (omit it for the original additive-only behaviour). When present, each key
is the name of a penalty row in the rubric's penalties block; the value is either a bare level
string or an object {"level": ..., "note": ...}. Accepted levels (case-insensitive, ES/EN):
    met       | cumple      -> no deduction
    partial   | parcial     -> 40% of the penalty magnitude (mirror of the 60% partial credit)
    not_met   | no_cumple   -> full penalty
Penalty magnitudes are read from the rubric (column E of the penalties block): a plain number
(e.g. "-15") deducts points; a percentage (e.g. "-100%") removes that share of the running total,
so "-100%" not-met zeroes the grade (invalid work). Point penalties apply first, then percentages.

Score labels must be exactly one of:
    "Meets Expectations" | "Partially Meets" | "Does Not Meet"
"""

import json
import sys
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

MULTIPLIERS = {
    "Meets Expectations": 1.0,
    "Partially Meets": 0.6,
    "Does Not Meet": 0.0,
}

VALID_LABELS = set(MULTIPLIERS.keys())

# Fraction of a penalty applied at each level (mirrors the additive 0.6 partial-credit rule).
PENALTY_LEVEL_FACTOR = {
    "met": 0.0, "cumple": 0.0,
    "partial": 0.4, "parcial": 0.4,
    "not_met": 1.0, "no_cumple": 1.0, "not met": 1.0,
}


def _fmt(n):
    """Format a number without trailing '.0' for tidy observation notes."""
    n = round(n, 2)
    return str(int(n)) if n == int(n) else str(n)


def read_rubric(rubric_path):
    """Parse the rubric xlsx into (criteria, penalties).

    criteria  -> list of (name, weight) for the additive block.
    penalties -> list of {"name", "text"} for the optional penalties block.

    The additive block is the leading rows whose Points column (E) is numeric. The penalties
    block, if present, starts at the bold heading row (column A begins with "Penal…") that
    rubric-creator emits, followed by a header row and one row per penalty whose column E holds
    free-text amounts like "-15" or "-100%". Rubrics with no penalties block just return [].
    """
    wb = load_workbook(rubric_path, data_only=True)
    ws = wb.active
    criteria, penalties = [], []
    in_penalties = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        pts = row[4] if len(row) >= 5 else None
        if name is None or str(name).strip() == "":
            continue  # blank separator row
        name = str(name).strip()
        low = name.lower()
        if low.startswith("penal"):          # penalties block heading -> everything below is penalties
            in_penalties = True
            continue
        if low in ("criterio", "criterion"):  # header row of either block
            continue
        if in_penalties:
            penalties.append({"name": name, "text": "" if pts is None else str(pts).strip()})
            continue
        if pts is None:
            continue
        try:
            weight = float(pts)
        except (TypeError, ValueError):
            # Non-numeric where a weight was expected: we've reached the penalties block without a
            # recognised heading. Treat this and the rest of the sheet as penalties.
            in_penalties = True
            penalties.append({"name": name, "text": str(pts).strip()})
            continue
        criteria.append((name, weight))
    return criteria, penalties


def parse_penalty_amount(text):
    """Parse a penalty amount into (kind, magnitude), or None if unparseable.

    kind == "percent" -> magnitude is a fraction (0..1+), e.g. "-100%" -> ("percent", 1.0)
    kind == "points"  -> magnitude is a point amount,   e.g. "-15"    -> ("points", 15.0)
    """
    if text is None:
        return None
    s = str(text).strip().replace("−", "-").replace(" ", "")  # normalise unicode minus
    if not s:
        return None
    is_pct = s.endswith("%")
    if is_pct:
        s = s[:-1]
    s = s.lstrip("+-")
    try:
        val = float(s)
    except ValueError:
        return None
    return ("percent", val / 100.0) if is_pct else ("points", val)


def apply_penalties(base_total, rubric_penalties, row_penalties):
    """Apply the row's penalty assessments to base_total. Returns (final_total, notes).

    Point penalties are subtracted first, then percentage penalties are applied multiplicatively,
    then the total is clamped at 0. notes is a list of human-readable strings for the
    observations cell, one per penalty that actually deducted points.
    """
    if not rubric_penalties or not row_penalties:
        return base_total, []

    point_deduction = 0.0
    pct_fractions = []
    notes = []
    for pen in rubric_penalties:
        entry = row_penalties.get(pen["name"])
        if entry is None:
            continue
        if isinstance(entry, dict):
            level = str(entry.get("level", "")).strip().lower()
            note = str(entry.get("note", "")).strip()
        else:
            level, note = str(entry).strip().lower(), ""
        factor = PENALTY_LEVEL_FACTOR.get(level, 0.0)
        if factor == 0.0:  # met / cumple / unknown-benign -> no deduction
            continue
        parsed = parse_penalty_amount(pen["text"])
        if parsed is None:
            continue
        kind, magnitude = parsed
        suffix = f" ({note})" if note else ""
        if kind == "points":
            ded = magnitude * factor
            point_deduction += ded
            notes.append(f"PENALTY — {pen['name']}: -{_fmt(ded)} pts [{level}]{suffix}")
        else:
            frac = min(1.0, magnitude * factor)
            pct_fractions.append(frac)
            notes.append(f"PENALTY — {pen['name']}: -{_fmt(frac * 100)}% [{level}]{suffix}")

    total = base_total - point_deduction
    for frac in pct_fractions:
        total *= (1.0 - frac)
    return max(0.0, total), notes


def build_header(criteria):
    headers = ["Name / Group"] + [c[0] for c in criteria] + ["Total", "Observations"]
    return headers


def create_grades_file(grades_path, criteria):
    wb = Workbook()
    ws = wb.active
    ws.title = "Grades"

    headers = build_header(criteria)
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Column widths
    ws.column_dimensions['A'].width = 28
    for i in range(len(criteria)):
        col_letter = ws.cell(1, i + 2).column_letter
        ws.column_dimensions[col_letter].width = 22
    # Total
    total_col = ws.cell(1, len(criteria) + 2).column_letter
    ws.column_dimensions[total_col].width = 10
    # Observations
    obs_col = ws.cell(1, len(criteria) + 3).column_letter
    ws.column_dimensions[obs_col].width = 80

    ws.freeze_panes = "A2"
    wb.save(grades_path)


def append_row(grades_path, criteria, penalties, row_data):
    wb = load_workbook(grades_path)
    ws = wb.active

    name = row_data["name"]
    scores = row_data["scores"]
    raw_obs = row_data.get("observations", "")
    if isinstance(raw_obs, dict):
        observations = " ".join(
            f"{k}: {v}" for k, v in raw_obs.items() if v
        )
    else:
        observations = raw_obs

    # Validate labels
    for cname, label in scores.items():
        if label not in VALID_LABELS:
            raise ValueError(f"Invalid score label '{label}' for criterion '{cname}'. Must be one of: {VALID_LABELS}")

    # Additive total, then apply the optional penalties block.
    base_total = sum(
        weight * MULTIPLIERS[scores.get(cname, "Does Not Meet")]
        for cname, weight in criteria
    )
    total, penalty_notes = apply_penalties(base_total, penalties, row_data.get("penalties", {}))
    if penalty_notes:
        observations = (observations + " " if observations else "") + " ".join(penalty_notes)

    row = [name] + [scores.get(cname, "Does Not Meet") for cname, _ in criteria] + [round(total, 1), observations]
    ws.append(row)

    data_row = ws.max_row
    num_cols = len(row)

    for col in range(1, num_cols + 1):
        cell = ws.cell(data_row, col)
        is_total = col == num_cols - 1
        kwargs = dict(wrap_text=True, vertical='top')
        if is_total:
            kwargs['horizontal'] = 'center'
            cell.font = Font(bold=True)
        cell.alignment = Alignment(**kwargs)

    wb.save(grades_path)
    print(f"Appended row for '{name}' → total: {round(total, 1)}")


def main():
    if len(sys.argv) != 4:
        print("Usage: python append_grade.py <rubric_xlsx> <grades_xlsx> <row_json>")
        sys.exit(1)

    rubric_path, grades_path, row_json_path = sys.argv[1], sys.argv[2], sys.argv[3]

    criteria, penalties = read_rubric(rubric_path)

    if not Path(grades_path).exists():
        create_grades_file(grades_path, criteria)
        print(f"Created grades file: {grades_path}")

    with open(row_json_path) as f:
        row_data = json.load(f)

    append_row(grades_path, criteria, penalties, row_data)


if __name__ == "__main__":
    main()
