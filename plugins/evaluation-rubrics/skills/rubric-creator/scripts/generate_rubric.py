"""
generate_rubric.py — Rubric Creator output script
Usage: python generate_rubric.py <input_json> <output_path>

input_json: path to a JSON file with the following structure:
{
  "language": "es" | "en",            # optional, default "es"
  "criteria": [                        # additive rubric — weights MUST sum to 100
    {
      "name": "...",
      "cumple": "...",
      "parcial": "...",
      "no_cumple": "...",
      "pts": <int>
    },
    ...
  ],
  "penalties": [                       # OPTIONAL — negatively-scored rows (see below)
    {
      "name": "...",
      "cumple": "...",
      "parcial": "...",                # optional, defaults to "N/A"
      "no_cumple": "...",
      "penalty": "-15" | "-100%" | "..."   # free text; key may also be "penalizacion"
    },
    ...
  ],
  "penalties_title": "..."             # OPTIONAL — override the penalties block heading
}

Penalties are OPTIONAL and OUT of the 100-point additive total: the sum==100 check
applies only to `criteria`. When present, they render as a labelled block below the
additive table (blank separator row → bold heading → header row → penalty rows).
They only ever DEDUCT points ("scored negatively"); they never add. Omitting the
field reproduces the original additive-only output exactly (backward compatible).
"""

import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

HEADERS = {
    "es": ["Criterio", "Cumple (100%)", "Cumple Parcialmente (60%)", "No Cumple (0%)", "Puntos"],
    "en": ["Criterion", "Meets Expectations (100%)", "Partially Meets (60%)", "Does Not Meet (0%)", "Points"],
}

SHEET_NAMES = {"es": "Rubrica", "en": "Rubric"}

COL_WIDTHS = [28, 52, 52, 52, 10]

# --- Optional penalties block (negatively-scored) ---
PENALTY_HEADERS = {
    "es": ["Criterio", "Cumple", "Cumple Parcialmente", "No Cumple", "Penalización"],
    "en": ["Criterion", "Meets", "Partially Meets", "Does Not Meet", "Penalty"],
}
PENALTY_TITLE = {
    "es": "Penalizaciones — se califican en negativo (solo bajan puntos si no se cumplen)",
    "en": "Penalties — scored negatively (deductions applied only when a row is not met)",
}
PENALTY_COL_E_WIDTH = 26  # penalty text needs more room than the numeric Points column


def _style_header_row(ws, row_idx):
    for cell in ws[row_idx]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical='top')


def generate(input_json_path, output_path):
    with open(input_json_path) as f:
        data = json.load(f)

    lang = data.get("language", "es")
    criteria = data["criteria"]
    penalties = data.get("penalties") or []

    assert sum(c["pts"] for c in criteria) == 100, \
        f"Points must sum to 100, got {sum(c['pts'] for c in criteria)}"

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAMES[lang]

    # --- Additive rubric (rows sum to 100) ---
    headers = HEADERS[lang]
    ws.append(headers)
    _style_header_row(ws, 1)

    for col_letter, width in zip("ABCDE", COL_WIDTHS):
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A2"

    for c in criteria:
        ws.append([c["name"], c["cumple"], c["parcial"], c["no_cumple"], c["pts"]])
        row = ws.max_row
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 5).font = Font(bold=True)
        for col in range(1, 6):
            kwargs = dict(wrap_text=True, vertical='top')
            if col == 5:
                kwargs['horizontal'] = 'center'
            ws.cell(row, col).alignment = Alignment(**kwargs)

    # --- Optional penalties block (out of the 100 total) ---
    if penalties:
        ws.column_dimensions["E"].width = PENALTY_COL_E_WIDTH

        ws.append([])  # blank separator row

        title = data.get("penalties_title") or PENALTY_TITLE[lang]
        ws.append([title])
        ws.cell(ws.max_row, 1).font = Font(bold=True)
        ws.cell(ws.max_row, 1).alignment = Alignment(wrap_text=True, vertical='top')

        ws.append(PENALTY_HEADERS[lang])
        _style_header_row(ws, ws.max_row)

        for p in penalties:
            penalty_text = p.get("penalty", p.get("penalizacion", ""))
            ws.append([
                p["name"],
                p.get("cumple", ""),
                p.get("parcial", "N/A"),
                p.get("no_cumple", ""),
                penalty_text,
            ])
            row = ws.max_row
            ws.cell(row, 1).font = Font(bold=True)
            for col in range(1, 6):
                ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical='top')

    wb.save(output_path)
    msg = f"Saved: {output_path} ({len(criteria)} criteria, {sum(c['pts'] for c in criteria)} pts"
    if penalties:
        msg += f", {len(penalties)} penalties"
    print(msg + ")")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python generate_rubric.py <input_json> <output_path>")
        sys.exit(1)
    generate(sys.argv[1], sys.argv[2])
