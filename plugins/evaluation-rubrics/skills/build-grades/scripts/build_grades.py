"""
build_grades.py — Build Grades output script
Rebuilds the grades xlsx from scratch using all row_*.json files in a folder.

Usage:
    python build_grades.py <rubric_xlsx> <grades_xlsx> <evaluations_dir>

Arguments:
    rubric_xlsx      Path to the rubric .xlsx produced by rubric-creator
    grades_xlsx      Path to the grades output file (always recreated from scratch)
    evaluations_dir Directory containing row_*.json evaluation files

Row JSON structure (observations may be a dict or a legacy plain string):
{
  "name": "Group 1",
  "scores": {
    "Criterion Name 1": "Meets Expectations",
    "Criterion Name 2": "Partially Meets",
    "Criterion Name 3": "Does Not Meet"
  },
  "observations": {
    "Criterion Name 1": "",
    "Criterion Name 2": "only 2 out of 10 elements completed.",
    "Criterion Name 3": "missing price and link."
  }
}
"""

import json
import sys
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

MULTIPLIERS = {
    "Meets Expectations": 1.0,
    "Partially Meets": 0.6,
    "Does Not Meet": 0.0,
}

VALID_LABELS = set(MULTIPLIERS.keys())


def read_rubric(rubric_path):
    """Returns list of (criterion_name, weight) tuples from the rubric xlsx."""
    wb = load_workbook(rubric_path, data_only=True)
    ws = wb.active
    criteria = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, _, _, _, pts = row[0], row[1], row[2], row[3], row[4]
        if name and pts is not None:
            criteria.append((str(name).strip(), float(pts)))
    return criteria


def create_grades_file(grades_path, criteria):
    wb = Workbook()
    ws = wb.active
    ws.title = "Grades"

    headers = ["Name / Group"] + [c[0] for c in criteria] + ["Total", "Observations"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    ws.column_dimensions['A'].width = 28
    for i in range(len(criteria)):
        col_letter = ws.cell(1, i + 2).column_letter
        ws.column_dimensions[col_letter].width = 22
    total_col = ws.cell(1, len(criteria) + 2).column_letter
    ws.column_dimensions[total_col].width = 10
    obs_col = ws.cell(1, len(criteria) + 3).column_letter
    ws.column_dimensions[obs_col].width = 80

    ws.freeze_panes = "A2"
    wb.save(grades_path)


def compile_observations(raw_obs):
    """Compile observations dict or string into a single xlsx-ready string."""
    if isinstance(raw_obs, dict):
        return " ".join(f"{k}: {v}" for k, v in raw_obs.items() if v)
    return raw_obs or ""


def append_row(grades_path, criteria, row_data):
    wb = load_workbook(grades_path)
    ws = wb.active

    name = row_data["name"]
    scores = row_data["scores"]
    observations = compile_observations(row_data.get("observations", ""))

    for cname, label in scores.items():
        if label not in VALID_LABELS:
            raise ValueError(
                f"Invalid score label '{label}' for criterion '{cname}'. "
                f"Must be one of: {VALID_LABELS}"
            )

    total = sum(
        weight * MULTIPLIERS[scores.get(cname, "Does Not Meet")]
        for cname, weight in criteria
    )

    row = [name] + [scores.get(cname, "Does Not Meet") for cname, _ in criteria] + [round(total, 1), observations]
    ws.append(row)

    data_row = ws.max_row
    num_cols = len(row)

    for col in range(1, num_cols + 1):
        cell = ws.cell(data_row, col)
        is_total = col == num_cols - 1
        kwargs = dict(wrap_text=True, vertical='top')
        if is_total:
            kwargs['horizontal'] = 'center'
            cell.font = Font(bold=True)
        cell.alignment = Alignment(**kwargs)

    wb.save(grades_path)
    return name, round(total, 1)


def main():
    if len(sys.argv) != 4:
        print("Usage: python build_grades.py <rubric_xlsx> <grades_xlsx> <evaluations_dir>")
        sys.exit(1)

    rubric_path, grades_path, evaluations_dir = sys.argv[1], sys.argv[2], sys.argv[3]

    criteria = read_rubric(rubric_path)

    json_files = sorted(Path(evaluations_dir).glob("row_*.json"))
    if not json_files:
        print(f"No row_*.json files found in {evaluations_dir}")
        sys.exit(1)

    # Always recreate from scratch
    create_grades_file(grades_path, criteria)
    print(f"Created grades file: {grades_path}")

    count = 0
    for json_file in json_files:
        with open(json_file) as f:
            row_data = json.load(f)
        name, total = append_row(grades_path, criteria, row_data)
        print(f"  [{count + 1}] '{name}' → total: {total}")
        count += 1

    print(f"\nRebuilt {count} row(s).")


if __name__ == "__main__":
    main()
